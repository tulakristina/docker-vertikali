from flask import Flask, render_template_string, request, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
import re

app = Flask(__name__)

with open("tuscias_ver.xlsx", "rb") as f:
    TUSCIAS_BYTES = f.read()
with open("Statistika_bendroji_bendra_sablonas.xlsx", "rb") as f:
    SABLONAS_BYTES = f.read()
with open("PASMULKINTA_SABLONAS_BE PALYGINIMO.xlsx", "rb") as f:
    PASMULKINTA_BYTES = f.read()


def safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def fmt_pct(cv, pv):
    if cv == 0 and pv == 0:
        return "0%"
    if pv == 0:
        return "100%" if cv > 0 else ""
    return f"{round((cv - pv) / pv * 100, 1)}%"


def parse_source_file(file_obj, filename):
    fn = filename.lower()
    if fn.endswith('.csv'):
        wb = Workbook()
        ws = wb.active
        text = file_obj.read().decode('utf-8')
        reader = csv.reader(io.StringIO(text))
        for ri, row in enumerate(reader, 1):
            for ci, val in enumerate(row, 1):
                if val.strip() == '':
                    ws.cell(row=ri, column=ci).value = None
                else:
                    try:
                        ws.cell(row=ri, column=ci).value = float(val)
                    except ValueError:
                        ws.cell(row=ri, column=ci).value = val
    else:
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
    return ws


def _raw(ws, row, col):
    v = ws.cell(row=row, column=col).value
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def extract_data_ver(ws):
    v1 = [
        [safe_int(ws.cell(row=row, column=col).value) for col in range(3, 11)]
        for row in range(2, 20)
    ]
    v2 = [
        [safe_int(ws.cell(row=row, column=col).value) for col in range(3, 11)]
        for row in range(20, 54)
    ]
    # Rows 13 and 27 are formula-based totals in the source XLSX (eil. 1+9 and eil. 13+14).
    # Calculate from raw floats to avoid zeros when cached formula values are absent,
    # and to avoid truncation errors from summing already-int'd values.
    raw1 = [[_raw(ws, r, c) for c in range(3, 11)] for r in range(2, 20)]
    raw2 = [[_raw(ws, r, c) for c in range(3, 11)] for r in range(20, 54)]
    v1[17] = [safe_int(raw1[0][j] + raw1[12][j]) for j in range(8)]   # type 13 = 1 + 9
    v2[33] = [safe_int(raw1[17][j] + raw2[0][j])  for j in range(8)]  # type 27 = 13 + 14
    return v1, v2


# Source rows 2-19 → comparison template base row (prev-year row of each 4-row group)
# Type 11 (Techninės bibliotekos) is in the template at rows 91-94 but not in the source.
PART1_MAPPING = [
     35,  # src 2  → type 1
     39,  # src 3  → type 2
     43,  # src 4  → type 3
     47,  # src 5  → type 4
     51,  # src 6  → type 5
     55,  # src 7  → type 5a
     59,  # src 8  → type 5b
     63,  # src 9  → type 5c
     67,  # src 10 → type 5d
     71,  # src 11 → type 6
     75,  # src 12 → type 7
     79,  # src 13 → type 8
     83,  # src 14 → type 9
     87,  # src 15 → type 10
     # 91 skipped — type 11 not in source
     95,  # src 16 → type 12
     99,  # src 17 → type 12a
    103,  # src 18 → type 12b
    107,  # src 19 → type 13 (Iš viso)
]

# Source rows 20-53 → comparison template base row (prev-year row of each 4-row group)
# Groups for types 22a, 23a, 23b, 23c are not in the source and left empty.
PART2_MAPPING = [
    118,  # src 20 → type 14
    122,  # src 21 → type 15
    126,  # src 22 → type 15a
    130,  # src 23 → type 15b
    134,  # src 24 → type 16
    138,  # src 25 → type 16a
    142,  # src 26 → type 16b
    146,  # src 27 → type 17
    150,  # src 28 → type 17a
    154,  # src 29 → type 17b
    158,  # src 30 → type 18
    162,  # src 31 → type 18a
    166,  # src 32 → type 18b
    170,  # src 33 → type 19
    174,  # src 34 → type 20
    178,  # src 35 → type 20a
    182,  # src 36 → type 20b
    186,  # src 37 → type 20c
    190,  # src 38 → type 20d
    194,  # src 39 → type 21
    198,  # src 40 → type 21a
    202,  # src 41 → type 21b
    206,  # src 42 → type 21c
    210,  # src 43 → type 21d
    214,  # src 44 → type 22
    222,  # src 45 → type 22b  (22a row 218 is skipped — not in source)
    226,  # src 46 → type 22c
    230,  # src 47 → type 22d
    234,  # src 48 → type 23
    250,  # src 49 → type 23d  (23a/23b/23c rows 238-249 skipped — not in source)
    254,  # src 50 → type 24
    258,  # src 51 → type 25
    262,  # src 52 → type 26
    266,  # src 53 → type 27
]


def write_standard_report(values_1, values_2):
    template_wb = load_workbook(io.BytesIO(TUSCIAS_BYTES))
    ws = template_wb.active

    start_row_1, start_col = 35, 13
    for i, row in enumerate(values_1):
        for j, value in enumerate(row):
            ws.cell(row=start_row_1 + i, column=start_col + j, value=value)

    start_row_2 = 60
    for i, row in enumerate(values_2):
        for j, value in enumerate(row):
            ws.cell(row=start_row_2 + i, column=start_col + j, value=value)

    return template_wb


def write_comparison_report(cur_v1, cur_v2, prev_v1, prev_v2, cur_year, prev_year):
    template_wb = load_workbook(io.BytesIO(SABLONAS_BYTES))
    ws = template_wb.active

    DATA_COL = 14  # first of 8 data columns (cols 14-21)

    # Update year labels in the template (cols 1-2, rows 35-269)
    for row in range(35, 270):
        for col in [1, 2]:
            cell = ws.cell(row=row, column=col)
            v = cell.value
            if v in (2024, 2024.0):
                cell.value = prev_year
            elif v in (2025, 2025.0):
                cell.value = cur_year
            elif isinstance(v, str):
                if '2024' in v:
                    cell.value = v.replace('2024', str(prev_year))
                elif '2025' in v:
                    cell.value = v.replace('2025', str(cur_year))

    # Part 1 — explicit mapping (skips type 11 which is in template but not in source)
    for src_idx, tmpl_base in enumerate(PART1_MAPPING):
        for j in range(8):
            pv = prev_v1[src_idx][j]
            cv = cur_v1[src_idx][j]
            ws.cell(row=tmpl_base,     column=DATA_COL + j).value = pv
            ws.cell(row=tmpl_base + 1, column=DATA_COL + j).value = cv
            ws.cell(row=tmpl_base + 2, column=DATA_COL + j).value = cv - pv
            pct = fmt_pct(cv, pv)
            if pct:
                ws.cell(row=tmpl_base + 3, column=DATA_COL + j).value = pct

    # Part 2 — explicit mapping (skips 22a, 23a, 23b, 23c which have no source data)
    for src_idx, tmpl_base in enumerate(PART2_MAPPING):
        for j in range(8):
            pv = prev_v2[src_idx][j]
            cv = cur_v2[src_idx][j]
            ws.cell(row=tmpl_base,     column=DATA_COL + j).value = pv
            ws.cell(row=tmpl_base + 1, column=DATA_COL + j).value = cv
            ws.cell(row=tmpl_base + 2, column=DATA_COL + j).value = cv - pv
            pct = fmt_pct(cv, pv)
            if pct:
                ws.cell(row=tmpl_base + 3, column=DATA_COL + j).value = pct

    # Normalise data cells: remove bold, centre-align (keep template size 9 and font name)
    data_font = Font(name='Times New Roman', size=9, bold=False)
    data_align = Alignment(horizontal='center', vertical='center')
    for tmpl_base in PART1_MAPPING + PART2_MAPPING:
        for row_offset in range(4):
            for j in range(8):
                cell = ws.cell(row=tmpl_base + row_offset, column=DATA_COL + j)
                cell.font = data_font
                cell.alignment = data_align

    # Clear yellow bgColor (FFFFFFCC) from specific template cells that have it
    no_fill = PatternFill(fill_type=None)
    for r in range(194, 198):
        ws.cell(row=r, column=21).fill = no_fill
    for r in (258, 259):
        ws.cell(row=r, column=2).fill = no_fill

    # Hide empty rows (types not present in source) — hiding preserves merged cells
    for r in list(range(91, 95)) + list(range(218, 222)) + list(range(238, 250)):
        ws.row_dimensions[r].hidden = True

    return template_wb


_SPECIAL_NAMES = {
    'neringos m.':  'neringos sav.',
    'palangos m.':  'palangos m. sav.',
}

def _norm(name):
    n = re.sub(r'\braj\.', 'r.', name.strip(), flags=re.IGNORECASE).lower()
    return _SPECIAL_NAMES.get(n, n)


def write_pasmulkinta_report(csv_files):
    wb = load_workbook(io.BytesIO(PASMULKINTA_BYTES))
    ws = wb["PAV"]

    # Unmerge any merged cells overlapping the data area (rows 12-143, cols 3-10)
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= 143 and mr.max_row >= 12 and mr.min_col <= 10 and mr.max_col >= 3:
            ws.unmerge_cells(str(mr))

    # Build map: normalised name -> row number (skip headers and totals)
    skip_kw = ('apskritis', 'apskr.', 'iš viso')
    row_map = {}
    for r in range(12, 93):
        v = ws.cell(row=r, column=2).value
        if v and not any(k in str(v).lower() for k in skip_kw):
            row_map[_norm(str(v))] = r

    # Aggregate data across all uploaded CSVs
    # keys: normalised savivaldybė; values: list per column field
    data = {}
    for raw_bytes in csv_files:
        text = raw_bytes.decode('utf-8-sig')
        for row in csv.DictReader(io.StringIO(text)):
            name = _norm(row['savivaldybe'])
            d = data.setdefault(name, [0] * 8)
            d[0] += int(float(row.get('bib_skaicius', 0) or 0))
            d[1] += int(float(row.get('fiziniai_apsilankymai', 0) or 0))
            d[2] += int(float(row.get('virtualus_apsilankymai', 0) or 0))
            d[3] += int(float(row.get('dok_fondas', 0) or 0))
            d[4] += int(float(row.get('dok_isduotis', 0) or 0))
            d[5] += int(float(row.get('prof_bibliotekininkai', 0) or 0))
            d[6] += int(float(row.get('bib_kompiuteriai', 0) or 0))
            d[7] += int(float(row.get('bib_internetas', 0) or 0))

    # Write to template — 0 for rows with no matching CSV data
    data_font  = Font(name='Times New Roman', size=9, bold=False)
    data_align = Alignment(horizontal='center', vertical='center')
    for norm_name, r in row_map.items():
        d = data.get(norm_name, [0] * 8)
        ws.cell(row=r, column=3).value  = d[0]   # bib_skaicius
        ws.cell(row=r, column=4).value  = d[1]   # fiziniai_apsilankymai
        ws.cell(row=r, column=5).value  = d[2]   # virtualus_apsilankymai
        ws.cell(row=r, column=6).value  = d[3]   # dok_fondas
        ws.cell(row=r, column=7).value  = d[4]   # dok_isduotis
        ws.cell(row=r, column=8).value  = d[5]   # prof_bibliotekininkai
        ws.cell(row=r, column=9).value  = d[6]   # bib_kompiuteriai
        ws.cell(row=r, column=10).value = d[7]   # bib_internetas
        for col in range(3, 11):
            cell = ws.cell(row=r, column=col)
            cell.font      = data_font
            cell.alignment = data_align

    # Remove all sheets except PAV
    for name in [s for s in wb.sheetnames if s != "PAV"]:
        del wb[name]

    return wb


HTML_TEMPLATE = """
<!doctype html>
<html lang="lt">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Vertikalios ataskaitos kepėja</title>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
      background: #f0f2f5;
      display: flex; justify-content: center; align-items: center;
      min-height: 100vh; padding: 20px;
    }
    .card {
      background: #fff; border-radius: 12px; padding: 40px;
      box-shadow: 0 2px 16px rgba(0,0,0,0.08);
      max-width: 560px; width: 100%; text-align: center;
    }
    h1 { font-size: 1.3rem; color: #1a1a2e; margin-bottom: 8px; }
    .subtitle { color: #666; font-size: 0.9rem; margin-bottom: 20px; }

    /* Toggle */
    .toggle-row {
      display: flex; align-items: center; justify-content: center;
      gap: 10px; margin-bottom: 24px;
    }
    .toggle { position: relative; display: inline-block; width: 44px; height: 24px; }
    .toggle input { opacity: 0; width: 0; height: 0; }
    .toggle-slider {
      position: absolute; cursor: pointer; inset: 0;
      background: #ccd1d9; border-radius: 24px; transition: .3s;
    }
    .toggle-slider:before {
      content: ""; position: absolute; width: 18px; height: 18px;
      left: 3px; bottom: 3px; background: #fff; border-radius: 50%; transition: .3s;
    }
    .toggle input:checked + .toggle-slider { background: #4a6cf7; }
    .toggle input:checked + .toggle-slider:before { transform: translateX(20px); }
    .toggle-label { color: #555; font-size: 0.88rem; cursor: pointer; user-select: none; }

    /* Year inputs */
    .year-inputs { display: flex; gap: 12px; margin-bottom: 16px; }
    .year-input { flex: 1; text-align: left; }
    .year-input label { display: block; font-size: 0.82rem; color: #666; margin-bottom: 4px; }
    .year-input input {
      width: 100%; padding: 8px 12px; border: 1px solid #d0d5dd;
      border-radius: 8px; font-size: 0.95rem; text-align: center;
    }
    .year-input input:focus { outline: none; border-color: #4a6cf7; }

    /* Upload label */
    .upload-label {
      font-size: 0.82rem; color: #666; margin-bottom: 8px;
      text-align: left; font-weight: 500;
    }
    .upload-section { margin-bottom: 12px; }

    /* Drop zone */
    .drop-zone {
      border: 2px dashed #c0c6d0; border-radius: 10px;
      padding: 32px 20px; cursor: pointer;
      transition: all 0.2s ease; background: #fafbfc;
    }
    .drop-zone:hover, .drop-zone.drag-over {
      border-color: #4a6cf7; background: #eef1ff;
    }
    .drop-zone svg { width: 40px; height: 40px; color: #4a6cf7; margin-bottom: 8px; }
    .drop-zone p { color: #555; font-size: 0.9rem; }
    .drop-zone .hint { color: #999; font-size: 0.78rem; margin-top: 4px; }
    .file-name {
      margin-top: 10px; padding: 8px 14px; background: #eef1ff;
      border-radius: 8px; color: #4a6cf7; font-size: 0.85rem;
      display: none; align-items: center; justify-content: center; gap: 8px;
    }
    .file-name .remove {
      cursor: pointer; color: #e74c3c; font-weight: bold; font-size: 1.1rem;
    }
    input[type="file"] { display: none; }
    button[type="submit"] {
      margin-top: 16px; padding: 12px 32px; font-size: 1rem;
      background: #4a6cf7; color: #fff; border: none; border-radius: 8px;
      cursor: pointer; transition: background 0.2s; width: 100%;
    }
    button[type="submit"]:hover { background: #3a5ce5; }
    button[type="submit"]:disabled { background: #b0b8c9; cursor: not-allowed; }
  </style>
</head>
<body>
  <div class="card">
    <h1>Vertikalios ataskaitos kepėja</h1>
    <p class="subtitle">Įkelkite duomenis iš VDA sistemos</p>
    <p style="margin-bottom:16px;font-size:0.85rem"><a href="/pasmulkinta" style="color:#4a6cf7;text-decoration:none">Pasmulkintos ataskaitos kepėja &rarr;</a></p>

    <div class="toggle-row">
      <label class="toggle">
        <input type="checkbox" id="compToggle">
        <span class="toggle-slider"></span>
      </label>
      <span class="toggle-label">Su palyginimu</span>
    </div>

    <form action="/" method="post" enctype="multipart/form-data" id="uploadForm">
      <input type="hidden" name="mode" id="modeInput" value="standard">

      <div class="year-inputs" id="yearInputs" style="display:none">
        <div class="year-input">
          <label for="curYear">Šių metų duomenys</label>
          <input type="number" name="current_year" id="curYear" value="2025" min="2000" max="2100">
        </div>
        <div class="year-input">
          <label for="prevYear">Praėjusių metų duomenys</label>
          <input type="number" name="previous_year" id="prevYear" value="2024" min="2000" max="2100">
        </div>
      </div>

      <!-- Primary file -->
      <div class="upload-section">
        <div class="upload-label" id="lbl1" style="display:none">Šių metų failas</div>
        <div class="drop-zone" id="dz1">
          <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"
               stroke="currentColor" stroke-width="1.5">
            <path stroke-linecap="round" stroke-linejoin="round"
                  d="M3 16.5v2.25A2.25 2.25 0 005.25 21h13.5A2.25 2.25 0
                     0021 18.75V16.5m-13.5-9L12 3m0 0l4.5 4.5M12 3v13.5"/>
          </svg>
          <p>Nutempkite failą čia arba paspauskite</p>
          <p class="hint">.xlsx arba .csv formatai</p>
        </div>
        <input type="file" name="source_file" id="fi1" accept=".xlsx,.csv" required>
        <div class="file-name" id="fn1">
          <span id="ft1"></span>
          <span class="remove" id="rm1">&times;</span>
        </div>
      </div>

      <!-- Previous year file (comparison only) -->
      <div class="upload-section" id="prevSection" style="display:none">
        <div class="upload-label">Praėjusių metų failas</div>
        <div class="drop-zone" id="dz2">
          <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"
               stroke="currentColor" stroke-width="1.5">
            <path stroke-linecap="round" stroke-linejoin="round"
                  d="M3 16.5v2.25A2.25 2.25 0 005.25 21h13.5A2.25 2.25 0
                     0021 18.75V16.5m-13.5-9L12 3m0 0l4.5 4.5M12 3v13.5"/>
          </svg>
          <p>Nutempkite failą čia arba paspauskite</p>
          <p class="hint">.xlsx arba .csv formatai</p>
        </div>
        <input type="file" name="previous_file" id="fi2" accept=".xlsx,.csv">
        <div class="file-name" id="fn2">
          <span id="ft2"></span>
          <span class="remove" id="rm2">&times;</span>
        </div>
      </div>

      <button type="submit" id="submitBtn" disabled>Atsisiųsti ataskaitą</button>
    </form>
  </div>

  <script>
    const compToggle  = document.getElementById('compToggle');
    const modeInput   = document.getElementById('modeInput');
    const yearInputs  = document.getElementById('yearInputs');
    const prevSection = document.getElementById('prevSection');
    const lbl1        = document.getElementById('lbl1');
    const submitBtn   = document.getElementById('submitBtn');
    const fi1 = document.getElementById('fi1');
    const fi2 = document.getElementById('fi2');

    function updateSubmit() {
      const comp = compToggle.checked;
      const has1 = fi1.files.length > 0;
      const has2 = fi2.files.length > 0;
      submitBtn.disabled = comp ? !(has1 && has2) : !has1;
    }

    compToggle.addEventListener('change', () => {
      const on = compToggle.checked;
      yearInputs.style.display  = on ? 'flex'  : 'none';
      prevSection.style.display = on ? 'block' : 'none';
      lbl1.style.display        = on ? 'block' : 'none';
      modeInput.value           = on ? 'comparison' : 'standard';
      submitBtn.textContent     = on ? 'Atsisiųsti ataskaitą su palyginimu'
                                     : 'Atsisiųsti ataskaitą';
      fi2.required = on;
      updateSubmit();
    });

    function setupDZ(zoneId, inputId, nameId, textId, rmId) {
      const zone  = document.getElementById(zoneId);
      const input = document.getElementById(inputId);
      const name  = document.getElementById(nameId);
      const text  = document.getElementById(textId);
      const rm    = document.getElementById(rmId);

      function show(f) { text.textContent = f.name; name.style.display = 'flex'; updateSubmit(); }
      function clear()  { input.value = ''; name.style.display = 'none'; updateSubmit(); }

      zone.addEventListener('click', () => input.click());
      input.addEventListener('change', () => { if (input.files.length) show(input.files[0]); });
      zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag-over'); });
      zone.addEventListener('dragleave', () => zone.classList.remove('drag-over'));
      zone.addEventListener('drop', e => {
        e.preventDefault(); zone.classList.remove('drag-over');
        const f = e.dataTransfer.files[0];
        if (f && (f.name.endsWith('.xlsx') || f.name.endsWith('.csv'))) {
          input.files = e.dataTransfer.files; show(f);
        }
      });
      rm.addEventListener('click', e => { e.stopPropagation(); clear(); });
    }
    setupDZ('dz1','fi1','fn1','ft1','rm1');
    setupDZ('dz2','fi2','fn2','ft2','rm2');
  </script>
</body>
</html>
"""

PASMULKINTA_TEMPLATE = """
<!doctype html>
<html lang="lt">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Pasmulkintos ataskaitos kepėja</title>
  <style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
      background: #f0f2f5;
      display: flex; justify-content: center; align-items: center;
      min-height: 100vh; padding: 20px;
    }
    .card {
      background: #fff; border-radius: 12px; padding: 40px;
      box-shadow: 0 2px 16px rgba(0,0,0,0.08);
      max-width: 560px; width: 100%; text-align: center;
    }
    h1 { font-size: 1.3rem; color: #1a1a2e; margin-bottom: 8px; }
    .subtitle { color: #666; font-size: 0.9rem; margin-bottom: 24px; }
    .nav { margin-bottom: 20px; }
    .nav a { color: #4a6cf7; font-size: 0.85rem; text-decoration: none; }
    .nav a:hover { text-decoration: underline; }
    .drop-zone {
      border: 2px dashed #c0c6d0; border-radius: 10px;
      padding: 32px 20px; cursor: pointer;
      transition: all 0.2s ease; background: #fafbfc; margin-bottom: 12px;
    }
    .drop-zone:hover, .drop-zone.drag-over {
      border-color: #4a6cf7; background: #eef1ff;
    }
    .drop-zone svg { width: 40px; height: 40px; color: #4a6cf7; margin-bottom: 8px; }
    .drop-zone p { color: #555; font-size: 0.9rem; }
    .drop-zone .hint { color: #999; font-size: 0.78rem; margin-top: 4px; }
    input[type="file"] { display: none; }
    .file-name {
      margin-top: 10px; padding: 8px 14px; background: #eef1ff;
      border-radius: 8px; color: #4a6cf7; font-size: 0.85rem;
      display: none; align-items: center; justify-content: center; gap: 8px;
    }
    .file-name .remove { cursor: pointer; color: #e74c3c; font-weight: bold; font-size: 1.1rem; }
    button[type="submit"] {
      margin-top: 4px; padding: 12px 32px; font-size: 1rem;
      background: #4a6cf7; color: #fff; border: none; border-radius: 8px;
      cursor: pointer; transition: background 0.2s; width: 100%;
    }
    button[type="submit"]:hover { background: #3a5ce5; }
    button[type="submit"]:disabled { background: #b0b8c9; cursor: not-allowed; }
  </style>
</head>
<body>
  <div class="card">
    <div class="nav"><a href="/">&larr; Vertikalios ataskaitos kepėja</a></div>
    <h1>Pasmulkintos ataskaitos kepėja</h1>
    <p class="subtitle">Įkelkite CSV failus iš VDA sistemos</p>

    <form action="/pasmulkinta" method="post" enctype="multipart/form-data" id="pForm">
      <div class="drop-zone" id="pDz">
        <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24"
             stroke="currentColor" stroke-width="1.5">
          <path stroke-linecap="round" stroke-linejoin="round"
                d="M3 16.5v2.25A2.25 2.25 0 005.25 21h13.5A2.25 2.25 0
                   0021 18.75V16.5m-13.5-9L12 3m0 0l4.5 4.5M12 3v13.5"/>
        </svg>
        <p>Nutempkite failą čia arba paspauskite</p>
        <p class="hint">.csv formatas</p>
      </div>
      <input type="file" name="csv_files" id="pInput" accept=".csv">
      <div class="file-name" id="pFileName">
        <span id="pFileText"></span>
        <span class="remove" id="pRemove">&times;</span>
      </div>
      <button type="submit" id="pSubmit" disabled>Atsisiųsti ataskaitą</button>
    </form>
  </div>

  <script>
    const dz     = document.getElementById('pDz');
    const input  = document.getElementById('pInput');
    const fname  = document.getElementById('pFileName');
    const ftext  = document.getElementById('pFileText');
    const rm     = document.getElementById('pRemove');
    const btn    = document.getElementById('pSubmit');

    function show(f) { ftext.textContent = f.name; fname.style.display = 'flex'; btn.disabled = false; }
    function clear()  { input.value = ''; fname.style.display = 'none'; btn.disabled = true; }

    dz.addEventListener('click', () => input.click());
    input.addEventListener('change', () => { if (input.files.length) show(input.files[0]); });
    rm.addEventListener('click', e => { e.stopPropagation(); clear(); });
    dz.addEventListener('dragover', e => { e.preventDefault(); dz.classList.add('drag-over'); });
    dz.addEventListener('dragleave', () => dz.classList.remove('drag-over'));
    dz.addEventListener('drop', e => {
      e.preventDefault(); dz.classList.remove('drag-over');
      const f = e.dataTransfer.files[0];
      if (f && f.name.endsWith('.csv')) { input.files = e.dataTransfer.files; show(f); }
    });

    document.getElementById('pForm').addEventListener('submit', () => {
      setTimeout(clear, 1000);
    });
  </script>
</body>
</html>
"""


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        mode = request.form.get("mode", "standard")

        if "source_file" not in request.files:
            return "Failas neįkeltas", 400

        src = request.files["source_file"]
        try:
            ws_cur = parse_source_file(src, src.filename)
            cur_v1, cur_v2 = extract_data_ver(ws_cur)
        except Exception as e:
            return f":( Nepavyko nuskaityti failo: {e}", 500

        if mode == "comparison":
            if "previous_file" not in request.files or request.files["previous_file"].filename == "":
                return "Praėjusių metų failas neįkeltas", 400
            prev_file = request.files["previous_file"]
            try:
                ws_prev = parse_source_file(prev_file, prev_file.filename)
                prev_v1, prev_v2 = extract_data_ver(ws_prev)
            except Exception as e:
                return f":( Nepavyko nuskaityti praėjusių metų failo: {e}", 500

            cur_year  = int(request.form.get("current_year",  2025))
            prev_year = int(request.form.get("previous_year", 2024))

            try:
                wb = write_comparison_report(cur_v1, cur_v2, prev_v1, prev_v2, cur_year, prev_year)
            except Exception as e:
                return f"Klaida pildant šabloną: {e}", 500

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(
                output,
                as_attachment=True,
                download_name="Vertikali_ataskaita_su_palyginimu.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # Standard (single-year) report
        try:
            wb = write_standard_report(cur_v1, cur_v2)
        except Exception as e:
            return f"Klaida pildant šabloną: {e}", 500

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="Vertikali_ataskaita.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return render_template_string(HTML_TEMPLATE)


@app.route("/pasmulkinta", methods=["GET", "POST"])
def pasmulkinta():
    if request.method == "POST":
        files = request.files.getlist("csv_files")
        if not files or all(f.filename == "" for f in files):
            return "Failai neįkelti", 400
        try:
            csv_bytes = [f.read() for f in files if f.filename.endswith(".csv")]
            wb = write_pasmulkinta_report(csv_bytes)
        except Exception as e:
            return f"Klaida pildant šabloną: {e}", 500
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="Pasmulkinta_ataskaita.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return render_template_string(PASMULKINTA_TEMPLATE)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8001)
